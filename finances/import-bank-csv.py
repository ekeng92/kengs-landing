#!/usr/bin/env python3
"""Categorize and import STR expenses from Robinhood CC CSV into the finance tracker."""

import csv
import json
import os
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl

CSV_PATH = "/Users/ekeng/Downloads/684d9305-02c2-4cb6-b639-1cfcad1c6ea2.csv"
TRACKER_PATH = os.path.join(os.path.dirname(__file__), "kengs-landing-finance-tracker.xlsx")

STR_LOCATIONS = ["FAIRFIELD", "CORSICANA"]

VENDOR_MAP = {
    "Home Depot": {"category": "Repairs & Maintenance", "property": "360"},
    "Lowe's": {"category": "Repairs & Maintenance", "property": "360", "note": "Dishwasher"},
    "Lowes": {"category": "Repairs & Maintenance", "property": "360", "note": "Dishwasher"},
    "Ace Home Center": {"category": "Repairs & Maintenance", "property": "360"},
    "Capps True Value": {"category": "Repairs & Maintenance", "property": "360"},
    "Gexa Energy": {"category": "Utilities", "property": "360", "note": "Electric"},
    "Maga Waste": {"category": "Utilities", "property": "360", "note": "Trash service"},
    "Brookshire Brothers": {"category": "Supplies", "property": "360", "note": "Property supplies"},
    "Cooper Farms": {"category": "Supplies", "property": "360"},
    "Brenco": {"category": "Repairs & Maintenance", "property": "360"},
    "Dfw Estate Liquidators": {"category": "Supplies", "property": "360", "note": "Furnishing"},
    "Intuit": {"category": "Professional Services", "property": "General", "note": "TurboTax (prorate STR %)"},
    "Discount Tire": {"category": "Travel / Mileage", "property": "360", "note": "Tire replacement from property trips"},
    "North Texas Tollway": {"category": "Travel / Mileage", "property": "360", "note": "Tolls for property trips"},
}

TRAVEL_MERCHANTS = [
    "Burger King", "Taco Bell", "Pizza Hut", "Braum's", "Wendy's",
    "McDonald's", "El Jimador", "Shell", "Gulf", "Love's", "Exxon",
    "Chevron", "Conoco", "Allsup's",
]

PERSONAL_MERCHANTS = [
    "Dr. Paul Daum", "Heights Dermatology", "LabCorp", "Livingspring Family",
    "SeaWorld", "Great Wolf Lodge", "Girl Scouts", "Walnut Ridge Baptist",
    "Sport Clips", "Smoke Villa", "Mr. Smoke", "Bookoo Smoke", "Exotic Smoke",
    "Smoke Lab", "comma.ai", "Cheeky Monkeys", "Play Park", "Evite", "Fooda",
    "Mansfield Sports", "Groupon", "Temu", "IHOP", "Chick-fil-A", "Goodwill",
    "Raising Cane's", "Chili's", "Dessert by MommaCakes", "Sonic",
    "eBay", "CVS", "Portillo's", "Chuy's", "King Dragon", "Pepper",
    "RFC GRAPEVINE", "Mo Bettahs", "Andy's Frozen", "Market Street",
    "Lamar Food", "Rockin S", "H-E-B", "Subway", "QT ",
    "Kroger", "7 Eleven", "St. John Lutheran", "Mansfield Mission",
]


def is_personal(merchant, desc):
    for pm in PERSONAL_MERCHANTS:
        if pm.lower() in merchant.lower() or pm.lower() in desc.lower():
            return True
    return False


def match_vendor(merchant):
    for vm, info in VENDOR_MAP.items():
        if vm.lower() in merchant.lower():
            return info
    return None


def is_travel_merchant(merchant):
    for tm in TRAVEL_MERCHANTS:
        if tm.lower() in merchant.lower():
            return True
    return False


def near_property(desc):
    return any(loc in desc.upper() for loc in STR_LOCATIONS)


def categorize_csv(csv_path):
    str_expenses = []

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["Type"] not in ("Purchase", "Refund"):
                continue

            merchant = row["Merchant"]
            desc = row.get("Description", "") or ""
            amt = float(row["Amount"])
            date = row["Date"]

            if is_personal(merchant, desc):
                continue

            is_near = near_property(desc)
            matched = match_vendor(merchant)
            is_travel = is_travel_merchant(merchant)

            if matched:
                expense = {
                    "date": date,
                    "merchant": merchant,
                    "amount": amt,
                    "category": matched["category"],
                    "property": matched.get("property", "360"),
                    "description": desc[:50] if desc else merchant,
                    "note": matched.get("note", ""),
                    "near_property": is_near,
                    "verify": not is_near and "Home Depot" in merchant,
                }
                # Handle refunds
                if row["Type"] == "Refund":
                    expense["description"] = f"REFUND: {expense['description']}"
                str_expenses.append(expense)

            elif is_travel and is_near:
                expense = {
                    "date": date,
                    "merchant": merchant,
                    "amount": amt,
                    "category": "Travel / Mileage",
                    "property": "360",
                    "description": f"Travel meals/fuel - {merchant}",
                    "note": "Trip to Fairfield",
                    "near_property": True,
                    "verify": False,
                }
                str_expenses.append(expense)

    str_expenses.sort(key=lambda x: x["date"])
    return str_expenses


def write_to_tracker(expenses, tracker_path):
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb["Expenses"]

    # Find next empty row
    next_row = None
    for r in range(2, 202):
        if ws.cell(row=r, column=1).value is None:
            next_row = r
            break

    if next_row is None:
        print("ERROR: Expenses sheet is full")
        return 0

    written = 0
    for exp in expenses:
        if next_row > 201:
            print(f"WARNING: Sheet full after {written} entries")
            break

        dt = datetime.strptime(exp["date"], "%Y-%m-%d")

        # A: Date
        ws.cell(row=next_row, column=1, value=dt)
        ws.cell(row=next_row, column=1).number_format = "YYYY-MM-DD"
        # B: Month — formula, don't touch
        # C: Category
        ws.cell(row=next_row, column=3, value=exp["category"])
        # D: Vendor/Payee
        ws.cell(row=next_row, column=4, value=exp["merchant"])
        # E: Description
        note_parts = [exp["description"]]
        if exp["note"]:
            note_parts.append(exp["note"])
        if exp["verify"]:
            note_parts.append("VERIFY: GP/Mansfield HD - STR or personal?")
        ws.cell(row=next_row, column=5, value=" | ".join(note_parts))
        # F: Amount (absolute value — refunds are negative in CSV)
        ws.cell(row=next_row, column=6, value=exp["amount"])
        ws.cell(row=next_row, column=6).number_format = "#,##0.00"
        # G: Payment Method
        ws.cell(row=next_row, column=7, value="Robinhood CC")
        # H: Receipt?
        ws.cell(row=next_row, column=8, value="N")

        next_row += 1
        written += 1

    wb.save(tracker_path)
    return written


def main():
    csv_path = CSV_PATH
    if len(sys.argv) >= 2:
        csv_path = sys.argv[1]

    print(f"Reading: {csv_path}")
    expenses = categorize_csv(csv_path)

    print(f"\nTotal STR expenses found: {len(expenses)}")
    total = sum(e["amount"] for e in expenses)
    verify_count = sum(1 for e in expenses if e["verify"])
    print(f"Net total: ${total:,.2f}")
    print(f"Flagged for verification: {verify_count}")

    by_cat = defaultdict(lambda: {"count": 0, "total": 0.0})
    for e in expenses:
        by_cat[e["category"]]["count"] += 1
        by_cat[e["category"]]["total"] += e["amount"]

    print("\nBy category:")
    for cat, d in sorted(by_cat.items(), key=lambda x: x[1]["total"], reverse=True):
        print(f"  {cat:30s}  {d['count']:>3} txns  ${d['total']:>8.2f}")

    print("\n--- Writing to tracker ---")
    written = write_to_tracker(expenses, TRACKER_PATH)
    print(f"\nDone! Wrote {written} expenses to tracker.")
    print(f"Tracker: {TRACKER_PATH}")

    if verify_count > 0:
        print(f"\n⚠️  {verify_count} entries tagged 'VERIFY' — review in Excel for Home Depot GP/Mansfield charges that may be personal.")


if __name__ == "__main__":
    main()
