#!/usr/bin/env python3
"""Import STR (360 County Road) expenses from Chase bank account CSVs.

Accounts:
  6113 — Property checking (mortgages for Ironwood/Marlow also here)
  3125 — Primary checking (some STR purchases via Zelle/FB Marketplace)
  1047 — Chase credit card (Starlink, some hardware stores)

Categorizations based on SAGE Keng's answers (Apr 13 2026 session).
"""

import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

TRACKER = '/Users/ekeng/IdeaProjects/kengs-landing/finances/kengs-landing-finance-tracker.xlsx'
YELLOW = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

def date_to_month(date_str):
    """Convert MM/DD/YYYY to 'Month YYYY'."""
    months = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']
    parts = date_str.split('/')
    return f"{months[int(parts[0])-1]} {parts[2]}"

# === STR (360 County Road) EXPENSES ===
# Format: (date, category, vendor, description, amount, payment_method, receipt, status)
STR_EXPENSES = [
    # ─── 6113: CONTRACTORS ─────────────────────────────────────────
    # Jeff/Jerry Yancey — Fairfield contractor
    ('12/16/2025', 'Repairs & maintenance', 'Jeff Yancey', 'Contractor - property work (Fairfield)', 510.00, 'Zelle (6113)', 'CC', 'Business'),
    ('12/24/2025', 'Repairs & maintenance', 'Jeff Yancey', 'Contractor - property work (Fairfield)', 700.00, 'Zelle (6113)', 'CC', 'Business'),
    ('01/12/2026', 'Repairs & maintenance', 'Jeff Yancey', 'Contractor - property work (Fairfield)', 775.00, 'Zelle (6113)', 'CC', 'Business'),
    ('02/09/2026', 'Repairs & maintenance', 'Jerry Yancey', 'Contractor - property work (Fairfield)', 1833.33, 'Zelle (6113)', 'CC', 'Business'),
    ('03/06/2026', 'Repairs & maintenance', 'Jerry Yancey', 'Contractor - small payment', 6.82, 'Zelle (6113)', 'CC', 'Business'),

    # Jacob Hance — contractor
    ('01/12/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 150.00, 'Zelle (6113)', 'CC', 'Business'),
    ('01/12/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 255.00, 'Zelle (6113)', 'CC', 'Business'),
    ('01/12/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 40.00, 'Zelle (6113)', 'CC', 'Business'),
    ('01/14/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 800.00, 'Zelle (6113)', 'CC', 'Business'),
    ('01/20/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 801.00, 'Zelle (6113)', 'CC', 'Business'),
    ('01/20/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 60.00, 'Zelle (6113)', 'CC', 'Business'),
    ('01/20/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 25.00, 'Zelle (6113)', 'CC', 'Business'),

    # Adrian Sanchez / Chance Stockley — new contractor
    ('03/23/2026', 'Repairs & maintenance', 'Adrian/Chance (contractor)', 'Contractor - property work', 400.00, 'Zelle (6113)', 'CC', 'Business'),
    ('03/23/2026', 'Repairs & maintenance', 'Adrian/Chance (contractor)', 'Contractor - property work', 52.91, 'Zelle (6113)', 'CC', 'Business'),
    ('03/24/2026', 'Repairs & maintenance', 'Adrian/Chance (contractor)', 'Contractor - property work', 210.00, 'Zelle (6113)', 'CC', 'Business'),
    ('03/31/2026', 'Repairs & maintenance', 'Adrian/Chance (contractor)', 'Contractor - property work', 400.00, 'Zelle (6113)', 'CC', 'Business'),
    ('04/01/2026', 'Repairs & maintenance', 'Adrian/Chance (contractor)', 'Contractor - property work', 400.00, 'Zelle (6113)', 'CC', 'Business'),

    # ─── 6113: UTILITIES ──────────────────────────────────────────
    ('02/24/2026', 'Utilities', 'MAGA Waste LLC', 'Trash service - 360 CR', 85.00, 'ACH (6113)', 'CC', 'Business'),
    ('03/17/2026', 'Utilities', 'MAGA Waste LLC', 'Trash service - 360 CR', 85.00, 'ACH (6113)', 'CC', 'Business'),
    ('01/20/2026', 'Utilities', 'Thompson Water', 'Water service - 360 CR', 100.00, 'BillPay (6113)', 'CC', 'Business'),
    ('03/09/2026', 'Utilities', 'Thompson Water', 'Water service - 360 CR', 101.00, 'BillPay (6113)', 'CC', 'Business'),

    # ─── 6113: FB MARKETPLACE (uncertain) ─────────────────────────
    ('04/01/2026', 'Supplies', 'Moises (FB Marketplace)', 'Furniture/item for STR (probable)', 60.00, 'Zelle (6113)', 'N', 'Review'),
    ('02/17/2026', 'Supplies', 'Santiago Medellin (FB Marketplace)', 'Furniture/item for STR (probable)', 50.00, 'Zelle (6113)', 'N', 'Review'),

    # ─── 3125: CONTRACTORS ────────────────────────────────────────
    # Jerry Yancey — startup phase (Nov 2025)
    ('11/17/2025', 'Repairs & maintenance', 'Jerry Yancey', 'Contractor - startup property work', 500.00, 'Zelle (3125)', 'CC', 'Business'),
    ('11/17/2025', 'Repairs & maintenance', 'Jerry Yancey', 'Contractor - startup property work', 500.00, 'Zelle (3125)', 'CC', 'Business'),

    # Jacob Hance — from primary checking
    ('01/07/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 100.00, 'Zelle (3125)', 'CC', 'Business'),
    ('01/08/2026', 'Repairs & maintenance', 'Jacob Hance', 'Contractor - property work', 145.00, 'Zelle (3125)', 'CC', 'Business'),

    # John Lomas — contractor
    ('02/09/2026', 'Repairs & maintenance', 'John Lomas', 'Contractor - property work', 75.00, 'Zelle (3125)', 'CC', 'Business'),
    ('02/11/2026', 'Repairs & maintenance', 'John Lomas', 'Contractor - property work', 275.00, 'Zelle (3125)', 'CC', 'Business'),

    # John Flooring
    ('12/10/2025', 'Repairs & maintenance', 'John Flooring', 'Flooring work - 360 CR', 300.00, 'Zelle (3125)', 'CC', 'Business'),

    # Adrian/Chance — also paid from 3125
    ('03/24/2026', 'Repairs & maintenance', 'Adrian/Chance (contractor)', 'Contractor - property work', 50.23, 'Zelle (3125)', 'CC', 'Business'),
    ('03/25/2026', 'Repairs & maintenance', 'Adrian/Chance (contractor)', 'Contractor - property work', 190.00, 'Zelle (3125)', 'CC', 'Business'),

    # ─── 3125: CLEANING ──────────────────────────────────────────
    # Julie Rose — cleaner for 360 CR
    # NOTE: Already in tracker from prior import. Will be deduped.
    ('03/30/2026', 'Cleaning & turnover', 'Julie Rose', 'Cleaning - 360 CR', 50.00, 'Zelle (3125)', 'CC', 'Business'),
    ('04/06/2026', 'Cleaning & turnover', 'Julie Rose', 'Cleaning - 360 CR', 60.00, 'Zelle (3125)', 'CC', 'Business'),
    ('04/08/2026', 'Cleaning & turnover', 'Julie Rose', 'Cleaning - 360 CR', 60.00, 'Zelle (3125)', 'CC', 'Business'),

    # ─── 3125: FURNITURE / SUPPLIES ──────────────────────────────
    ('02/17/2026', 'Supplies', 'FB Marketplace - Sectionals', 'Sectional sofa for STR', 300.00, 'Zelle (3125)', 'N', 'Business'),
    ('02/17/2026', 'Supplies', 'FB Marketplace - Sectionals', 'Sectional sofa for STR (2nd)', 100.00, 'Zelle (3125)', 'N', 'Business'),
    ('02/17/2026', 'Supplies', 'FB Marketplace - Hutch', 'Hutch for STR', 125.00, 'Zelle (3125)', 'N', 'Business'),
    ('02/17/2026', 'Supplies', 'FB Marketplace - Chairs', 'Chairs for STR', 25.00, 'Zelle (3125)', 'N', 'Business'),
    ('02/17/2026', 'Supplies', 'FB Marketplace - Patio Chairs', 'Patio chairs for STR', 20.00, 'Zelle (3125)', 'N', 'Business'),
    ('02/17/2026', 'Supplies', 'FB Marketplace - Caulk', 'Caulk/supplies for STR', 16.00, 'Zelle (3125)', 'N', 'Business'),
    ('12/22/2025', 'Supplies', 'FB Marketplace - Whirlpool', 'Whirlpool appliance for STR', 20.00, 'Zelle (3125)', 'N', 'Business'),
    ('12/24/2025', 'Supplies', 'FB Marketplace - Whirlpool', 'Whirlpool appliance for STR', 230.00, 'Zelle (3125)', 'N', 'Business'),
    ('12/24/2025', 'Supplies', 'FB Marketplace - Dyson', 'Dyson vacuum for STR', 120.00, 'Zelle (3125)', 'N', 'Business'),
    ('03/02/2026', 'Supplies', 'Keng Bed', 'Bed for STR', 200.00, 'Zelle (3125)', 'N', 'Business'),

    # FB Marketplace — uncertain identity
    ('04/06/2026', 'Supplies', 'Brandon Conway (FB Marketplace)', 'Furniture for STR (probable)', 100.00, 'Zelle (3125)', 'N', 'Review'),
    ('02/05/2026', 'Supplies', 'John Neel (FB Marketplace)', 'Furniture for STR (probable)', 25.00, 'Zelle (3125)', 'N', 'Review'),
    ('02/12/2026', 'Supplies', 'John Neel (FB Marketplace)', 'Furniture for STR (probable)', 25.00, 'Zelle (3125)', 'N', 'Review'),
    ('01/20/2026', 'Supplies', 'Lindsey McCoy (FB Marketplace)', 'Furniture for STR (probable)', 70.00, 'Zelle (3125)', 'N', 'Review'),
    ('02/25/2026', 'Supplies', 'FB Frozen Jeep Mobile', 'Item for STR (probable)', 20.00, 'Zelle (3125)', 'N', 'Review'),

    # ─── 3125: PROPERTY / UTILITY ────────────────────────────────
    ('02/10/2026', 'Supplies', 'Ez Go Wheels', 'Golf cart for STR property', 140.00, 'Zelle (3125)', 'N', 'Business'),
    ('02/23/2026', 'Utilities', 'Brio Water', 'Water delivery - STR', 70.00, 'Zelle (3125)', 'N', 'Business'),

    # ─── 3125: LARGE STARTUP (Nov 2025) ──────────────────────────
    ('11/24/2025', 'Repairs & maintenance', 'Cash withdrawal', 'Cash for Jeff Yancey + skid steer for 360 CR land clearing — need to determine split', 7000.00, 'Cash (3125)', 'N', 'Review'),
    ('11/24/2025', 'Repairs & maintenance', 'ATM Cash', 'Cash — probably Jeff Yancey payment', 500.00, 'Cash (3125)', 'N', 'Review'),

    # ─── 1047 CC: STARLINK ────────────────────────────────────────
    ('10/12/2025', 'Utilities', 'Starlink', 'Starlink dish/equipment purchase', 455.81, 'Chase CC (1047)', 'CC', 'Business'),
    ('10/21/2025', 'Utilities', 'Starlink', 'Starlink refund/adjustment', -63.75, 'Chase CC (1047)', 'CC', 'Business'),
    ('10/17/2025', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 120.00, 'Chase CC (1047)', 'CC', 'Business'),
    ('11/16/2025', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 44.45, 'Chase CC (1047)', 'CC', 'Business'),
    ('11/16/2025', 'Utilities', 'Starlink', 'Starlink Internet - additional charge', 5.00, 'Chase CC (1047)', 'CC', 'Business'),
    ('12/16/2025', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 50.00, 'Chase CC (1047)', 'CC', 'Business'),
    ('01/16/2026', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 50.00, 'Chase CC (1047)', 'CC', 'Business'),
    ('02/11/2026', 'Utilities', 'Starlink', 'Starlink Internet - adjustment', 11.30, 'Chase CC (1047)', 'CC', 'Business'),
    ('02/16/2026', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 50.00, 'Chase CC (1047)', 'CC', 'Business'),
    ('02/27/2026', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 41.04, 'Chase CC (1047)', 'CC', 'Business'),
    ('03/16/2026', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 120.00, 'Chase CC (1047)', 'CC', 'Business'),
    ('04/03/2026', 'Utilities', 'Starlink', 'Starlink Internet - monthly', 47.58, 'Chase CC (1047)', 'CC', 'Business'),

    # ─── 1047 CC: HOME/HARDWARE STORES ────────────────────────────
    ('10/14/2025', 'Repairs & maintenance', 'Home Depot #6817', 'Repairs/supplies (Chase CC, Oct pre-rental)', 186.78, 'Chase CC (1047)', 'CC', 'Review'),
    ('10/28/2025', 'Repairs & maintenance', "Lowe's #01511", 'Repairs/supplies (Chase CC, Oct)', 18.91, 'Chase CC (1047)', 'CC', 'Review'),
]

# === REVENUE: McMeen group 105-day rental ===
# Noah Parrish contract — 360 County Road
# 4 tenants paying via Zelle to 6113 on 04/07/2026
MCMEEN_GROUP_BOOKING = {
    'month': 'April 2026',
    'platform': 'Direct (Zelle)',
    'guest': 'Noah Parrish group (McMeen/Parrish/Armstrong/Stevens)',
    'checkin': '',   # Unknown exact dates
    'checkout': '',
    'nights': 105,
    'rate': 0,   # Will compute from total
    'cleaning': 0,
    'payout': 1051.50,
    'status': 'Active',
    'notes': '105-day rental. 4 tenants: Robert McMeen $262.50, Noah Parrish $263, Dylan Armstrong $263, Carson Stevens $263. Payments received 04/07/2026.',
}


def main():
    wb = openpyxl.load_workbook(TRACKER)
    ws_exp = wb['Expenses']
    ws_bk = wb['Bookings']

    # Build existing expense fingerprints for dedup: (date_str, vendor_lower, amount)
    existing = set()
    for r in range(2, 500):
        date_val = ws_exp.cell(row=r, column=1).value
        if date_val is None:
            break
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%m/%d/%Y')
        else:
            date_str = str(date_val)
        vendor = str(ws_exp.cell(row=r, column=4).value or '').lower().strip()
        amount = ws_exp.cell(row=r, column=6).value or 0
        existing.add((date_str, vendor, round(float(amount), 2)))
    existing_count = r - 2

    # Import expenses
    added = 0
    skipped = 0
    review_count = 0
    row = existing_count + 2  # Next empty row

    for (date, cat, vendor, desc, amount, payment, receipt, status) in STR_EXPENSES:
        # Dedup check
        fingerprint = (date, vendor.lower().strip(), round(amount, 2))
        if fingerprint in existing:
            skipped += 1
            continue

        month = date_to_month(date)
        ws_exp.cell(row=row, column=1, value=datetime.strptime(date, '%m/%d/%Y'))
        ws_exp.cell(row=row, column=2, value=month)
        ws_exp.cell(row=row, column=3, value=cat)
        ws_exp.cell(row=row, column=4, value=vendor)
        ws_exp.cell(row=row, column=5, value=desc)
        ws_exp.cell(row=row, column=6, value=amount)
        ws_exp.cell(row=row, column=7, value=payment)
        ws_exp.cell(row=row, column=8, value=receipt)
        ws_exp.cell(row=row, column=9, value=status)

        if status == 'Review':
            review_count += 1
            for c in range(1, 10):
                ws_exp.cell(row=row, column=c).fill = YELLOW

        existing.add(fingerprint)
        added += 1
        row += 1

    # Add McMeen group booking
    bk = MCMEEN_GROUP_BOOKING
    bk_row = 2
    while ws_bk.cell(row=bk_row, column=1).value is not None:
        bk_row += 1
    ws_bk.cell(row=bk_row, column=1, value=bk['month'])
    ws_bk.cell(row=bk_row, column=2, value=bk['platform'])
    ws_bk.cell(row=bk_row, column=3, value=bk['guest'])
    ws_bk.cell(row=bk_row, column=4, value=bk['checkin'])
    ws_bk.cell(row=bk_row, column=5, value=bk['checkout'])
    ws_bk.cell(row=bk_row, column=6, value=bk['nights'])
    ws_bk.cell(row=bk_row, column=7, value=bk['rate'])
    ws_bk.cell(row=bk_row, column=8, value=bk['cleaning'])
    ws_bk.cell(row=bk_row, column=9, value=bk['payout'])
    ws_bk.cell(row=bk_row, column=10, value=bk['status'])
    ws_bk.cell(row=bk_row, column=11, value=bk['notes'])

    wb.save(TRACKER)

    # Summary
    print(f'Chase Account Import Complete')
    print(f'{"="*50}')
    print(f'  Existing expenses: {existing_count}')
    print(f'  New expenses added: {added}')
    print(f'  Duplicates skipped: {skipped}')
    print(f'  Items needing review: {review_count}')
    print(f'  Total expenses now: {existing_count + added}')
    print()

    # Breakdown by category
    from collections import defaultdict
    by_cat = defaultdict(float)
    by_source = defaultdict(float)
    for (date, cat, vendor, desc, amount, payment, receipt, status) in STR_EXPENSES:
        fp = (date, vendor.lower().strip(), round(amount, 2))
        # Only count what was actually added
        by_cat[cat] += amount
        source = payment.split('(')[1].rstrip(')') if '(' in payment else payment
        by_source[source] += amount

    print('  By Category (all attempted):')
    for cat in sorted(by_cat, key=lambda c: -by_cat[c]):
        print(f'    {cat:30s}  ${by_cat[cat]:>10,.2f}')
    total = sum(by_cat.values())
    print(f'    {"TOTAL":30s}  ${total:>10,.2f}')

    print()
    print('  By Account Source:')
    for src in sorted(by_source, key=lambda s: -by_source[s]):
        print(f'    {src:20s}  ${by_source[src]:>10,.2f}')

    print()
    print('  McMeen group booking added (105-day rental, $1,051.50)')
    print()
    print('  ⚠ REMINDERS (not in any CSV — add manually when docs available):')
    print('    • 2 golf carts ~$2,400 (bill of sale exists)')
    print('    • Skid steer purchase (part of $7K cash withdrawal, exact cost TBD)')
    print('    • Amazon orders — defer to CSV match (Option C)')
    print(f'\n  Tracker: {TRACKER}')


if __name__ == '__main__':
    main()
