#!/usr/bin/env python3
"""
Import Airbnb CSV export into the Keng's Landing finance tracker.

Usage:
    python3 finances/import-airbnb-csv.py [path-to-airbnb-csv]

If no path is given, auto-detects the newest airbnb*.csv in ~/Downloads/.

Reads Airbnb transaction CSV, extracts Reservation rows, deduplicates against
existing bookings (by confirmation code or net payout + month match), and
writes new/enriched entries to the Bookings sheet.
"""

import csv
import glob
import sys
import os
from datetime import datetime

import openpyxl

# --- Configuration ---
TRACKER_PATH = os.path.join(os.path.dirname(__file__), "kengs-landing-finance-tracker.xlsx")
DOWNLOADS_DIR = os.path.expanduser("~/Downloads")
SHEET_NAME = "Bookings"
MAX_ROWS = 101  # Header + 100 data rows

# Bookings column indices (1-based)
COL_MONTH = 1       # A: "MMM YYYY"
COL_PLATFORM = 2    # B: "Airbnb"
COL_GUEST = 3       # C: Guest Name
COL_CHECKIN = 4     # D: Check-In date
COL_CHECKOUT = 5    # E: Check-Out date
COL_NIGHTS = 6      # F: Nights
COL_RATE = 7        # G: Nightly Rate
COL_CLEANING = 8    # H: Cleaning Fee
COL_GROSS = 9       # I: Gross Revenue
COL_FEES = 10       # J: Platform Fees
COL_NET = 11        # K: Net Payout
COL_NOTES = 12      # L: Notes


def parse_date(date_str):
    """Parse MM/DD/YYYY date string from Airbnb CSV."""
    if not date_str:
        return None
    return datetime.strptime(date_str.strip(), "%m/%d/%Y")


def to_month_str(dt):
    """Convert datetime to 'MMM YYYY' format matching the tracker."""
    return dt.strftime("%b %Y")


def parse_float(val):
    """Parse a numeric string, returning 0.0 for empty/None."""
    if not val or val.strip() == "":
        return 0.0
    return float(val.strip().replace(",", ""))


def read_csv_reservations(csv_path):
    """Read Airbnb CSV and return only Reservation rows as dicts."""
    reservations = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Type", "").strip() == "Reservation":
                reservations.append(row)
    return reservations


def find_existing_bookings(ws):
    """Scan the Bookings sheet and return existing entries for dedup.

    Returns:
        confirmation_codes: set of confirmation codes found in Notes column
        payout_month_map: dict of (net_payout_float, month_str) -> row_number
            for stub entries (no confirmation code yet)
        next_empty_row: first row with no data in column A
    """
    confirmation_codes = set()
    payout_month_map = {}
    next_empty_row = None

    for r in range(2, MAX_ROWS + 1):
        month_val = ws.cell(row=r, column=COL_MONTH).value
        if month_val is None:
            if next_empty_row is None:
                next_empty_row = r
            continue

        notes = ws.cell(row=r, column=COL_NOTES).value or ""
        net_val = ws.cell(row=r, column=COL_NET).value

        # Extract confirmation codes from notes (format: "Airbnb HMXXXXXXXX")
        for word in notes.split():
            if word.startswith("HM") and len(word) >= 8:
                confirmation_codes.add(word)

        # Track stub entries by payout amount + month for enrichment
        guest = ws.cell(row=r, column=COL_GUEST).value
        if net_val and not guest:
            # Stub entry — no guest name means it wasn't fully populated
            try:
                key = (round(float(net_val), 2), str(month_val).strip())
                payout_month_map[key] = r
            except (ValueError, TypeError):
                pass

    if next_empty_row is None:
        next_empty_row = MAX_ROWS + 1

    return confirmation_codes, payout_month_map, next_empty_row


def write_booking(ws, row_num, res):
    """Write a single reservation to a row in the Bookings sheet."""
    start_date = parse_date(res.get("Start date"))
    end_date = parse_date(res.get("End date"))
    nights = int(res["Nights"]) if res.get("Nights") else None
    gross = parse_float(res.get("Gross earnings"))
    service_fee = parse_float(res.get("Service fee"))
    cleaning_fee = parse_float(res.get("Cleaning fee"))
    pet_fee = parse_float(res.get("Pet fee"))
    amount = parse_float(res.get("Amount"))
    conf_code = res.get("Confirmation code", "").strip()

    # Net payout = Amount (what actually hits the bank)
    net_payout = amount if amount else gross - service_fee

    # Nightly rate = (Gross - Cleaning - Pet) / Nights
    nightly_rate = None
    if nights and nights > 0:
        nightly_rate = round((gross - cleaning_fee - pet_fee) / nights, 2)

    # Month from check-in date
    month_str = to_month_str(start_date) if start_date else ""

    ws.cell(row=row_num, column=COL_MONTH, value=month_str)
    ws.cell(row=row_num, column=COL_PLATFORM, value="Airbnb")
    ws.cell(row=row_num, column=COL_GUEST, value=res.get("Guest", "").strip())

    if start_date:
        ws.cell(row=row_num, column=COL_CHECKIN, value=start_date)
        ws.cell(row=row_num, column=COL_CHECKIN).number_format = "YYYY-MM-DD"
    if end_date:
        ws.cell(row=row_num, column=COL_CHECKOUT, value=end_date)
        ws.cell(row=row_num, column=COL_CHECKOUT).number_format = "YYYY-MM-DD"

    if nights:
        ws.cell(row=row_num, column=COL_NIGHTS, value=nights)
    if nightly_rate:
        ws.cell(row=row_num, column=COL_RATE, value=nightly_rate)
        ws.cell(row=row_num, column=COL_RATE).number_format = "#,##0.00"
    if cleaning_fee:
        ws.cell(row=row_num, column=COL_CLEANING, value=cleaning_fee)
        ws.cell(row=row_num, column=COL_CLEANING).number_format = "#,##0.00"

    ws.cell(row=row_num, column=COL_GROSS, value=gross)
    ws.cell(row=row_num, column=COL_GROSS).number_format = "#,##0.00"
    ws.cell(row=row_num, column=COL_FEES, value=service_fee)
    ws.cell(row=row_num, column=COL_FEES).number_format = "#,##0.00"
    ws.cell(row=row_num, column=COL_NET, value=net_payout)
    ws.cell(row=row_num, column=COL_NET).number_format = "#,##0.00"

    # Notes: confirmation code + any extras
    note_parts = [f"Airbnb {conf_code}"]
    if pet_fee > 0:
        note_parts.append(f"Pet fee: ${pet_fee:.2f}")
    tax = parse_float(res.get("Airbnb remitted tax"))
    if tax > 0:
        note_parts.append(f"Tax remitted: ${tax:.2f}")
    ws.cell(row=row_num, column=COL_NOTES, value=" | ".join(note_parts))


def find_latest_airbnb_csv():
    """Find the newest airbnb*.csv file in ~/Downloads/."""
    pattern = os.path.join(DOWNLOADS_DIR, "airbnb*.csv")
    matches = glob.glob(pattern, recursive=False)
    if not matches:
        # Also try uppercase
        matches = glob.glob(os.path.join(DOWNLOADS_DIR, "Airbnb*.csv"), recursive=False)
    if not matches:
        return None
    # Sort by modification time, newest first
    matches.sort(key=os.path.getmtime, reverse=True)
    return matches[0]


def main():
    if len(sys.argv) >= 2:
        csv_path = sys.argv[1]
    else:
        csv_path = find_latest_airbnb_csv()
        if csv_path:
            print(f"Auto-detected: {csv_path}")
            print(f"  (modified: {datetime.fromtimestamp(os.path.getmtime(csv_path)).strftime('%Y-%m-%d %H:%M')})")
        else:
            print("No airbnb*.csv found in ~/Downloads/.")
            print("Usage: python3 finances/import-airbnb-csv.py [path-to-airbnb-csv]")
            sys.exit(1)

    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found: {csv_path}")
        sys.exit(1)

    if not os.path.exists(TRACKER_PATH):
        print(f"Error: Finance tracker not found: {TRACKER_PATH}")
        sys.exit(1)

    # --- Read CSV ---
    reservations = read_csv_reservations(csv_path)
    print(f"Found {len(reservations)} reservation(s) in CSV")

    if not reservations:
        print("No reservations to import.")
        return

    # --- Open tracker ---
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb[SHEET_NAME]

    existing_codes, stub_map, next_row = find_existing_bookings(ws)
    print(f"Existing tracker: {len(existing_codes)} confirmation code(s), {len(stub_map)} stub(s), next empty row: {next_row}")

    added = 0
    enriched = 0
    skipped = 0

    for res in reservations:
        conf_code = res.get("Confirmation code", "").strip()
        amount = parse_float(res.get("Amount"))
        start_date = parse_date(res.get("Start date"))
        month_str = to_month_str(start_date) if start_date else ""

        # Dedup 1: Already imported (confirmation code in notes)
        if conf_code in existing_codes:
            print(f"  SKIP (already imported): {conf_code} — {res.get('Guest', '').strip()}")
            skipped += 1
            continue

        # Dedup 2: Stub entry exists (same net payout + month, no guest)
        stub_key = (round(amount, 2), month_str)
        stub_row = stub_map.get(stub_key)
        if stub_row:
            print(f"  ENRICH row {stub_row}: {conf_code} — {res.get('Guest', '').strip()} (${amount:.2f})")
            write_booking(ws, stub_row, res)
            # Remove from stub map so it's not matched again
            del stub_map[stub_key]
            enriched += 1
            continue

        # New entry
        if next_row > MAX_ROWS:
            print(f"  ERROR: Bookings sheet is full (max {MAX_ROWS - 1} rows). Cannot add more.")
            break

        print(f"  ADD row {next_row}: {conf_code} — {res.get('Guest', '').strip()} (${amount:.2f})")
        write_booking(ws, next_row, res)
        next_row += 1
        added += 1

    # --- Save ---
    wb.save(TRACKER_PATH)
    print(f"\nDone! Added: {added}, Enriched: {enriched}, Skipped: {skipped}")
    print(f"Tracker saved: {TRACKER_PATH}")


if __name__ == "__main__":
    main()
