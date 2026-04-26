"""
Generate the Keng's Landing STR Finance Tracker Excel workbook.
Run: python3 finances/generate-tracker.py
Output: finances/kengs-landing-finance-tracker.xlsx

Preserves existing data if the file already exists.
Uses SUMIFS/COUNTIFS for cross-sheet aggregation (no SUMPRODUCT array issues).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUTPUT_PATH = "/Users/ekeng/IdeaProjects/kengs-landing/finances/kengs-landing-finance-tracker.xlsx"

BOOKING_ROWS = 100  # rows 2-101
EXPENSE_ROWS = 200  # rows 2-201
BK_LAST = 1 + BOOKING_ROWS   # 101
EX_LAST = 1 + EXPENSE_ROWS   # 201

CATEGORIES = [
    "Mortgage Interest", "Property Taxes", "Insurance",
    "Repairs & Maintenance", "Supplies", "Utilities",
    "Cleaning & Turnover", "Platform Fees", "Professional Services",
    "Advertising", "Travel / Mileage", "Depreciation", "Other"
]

MONTHS_2026 = [
    "Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026", "Jun 2026",
    "Jul 2026", "Aug 2026", "Sep 2026", "Oct 2026", "Nov 2026", "Dec 2026"
]
DAYS_IN_MONTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

# ── Read existing data ──────────────────────────────────────────────────
existing_bookings = []
existing_expenses = []
existing_budgets = {}  # category -> annual budget amount

if os.path.exists(OUTPUT_PATH):
    old_wb = openpyxl.load_workbook(OUTPUT_PATH, data_only=False)

    if 'Bookings' in old_wb.sheetnames:
        ws = old_wb['Bookings']
        for row in range(2, ws.max_row + 1):
            vals = [ws.cell(row=row, column=c).value for c in range(1, 13)]
            if vals[0] is not None and not str(vals[0]).startswith('='):
                existing_bookings.append(vals)

    if 'Expenses' in old_wb.sheetnames:
        ws = old_wb['Expenses']
        for row in range(2, ws.max_row + 1):
            vals = [ws.cell(row=row, column=c).value for c in range(1, 10)]
            if vals[0] is not None and not str(vals[0]).startswith('='):
                existing_expenses.append(vals)

    if 'Budget vs Actual' in old_wb.sheetnames:
        ws = old_wb['Budget vs Actual']
        for row in range(5, 5 + len(CATEGORIES)):
            cat = ws.cell(row=row, column=1).value
            budget = ws.cell(row=row, column=2).value
            if cat and budget is not None and isinstance(budget, (int, float)):
                existing_budgets[cat] = budget

    old_wb.close()
    print(f"Preserved {len(existing_bookings)} bookings, {len(existing_expenses)} expenses, {len(existing_budgets)} budget entries")

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="D9E2D9", end_color="D9E2D9", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="2F5233")
BOLD = Font(bold=True)
BOLD_GREEN = Font(bold=True, size=12, color="2F5233")
MONEY_FMT = '#,##0.00'
PCT_FMT = '0.0%'
DATE_FMT = 'YYYY-MM-DD'
BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = BORDER


def style_area(ws, r1, r2, ncols):
    for r in range(r1, r2 + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER


def auto_width(ws, ncols, minw=12):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        maxw = minw
        for cell in ws[letter]:
            if cell.value:
                maxw = max(maxw, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(maxw, 28)


wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════
# SHEET 1: Bookings
# ═══════════════════════════════════════════════════════════════════════
ws_bk = wb.active
ws_bk.title = "Bookings"
ws_bk.sheet_properties.tabColor = "2F5233"

bk_headers = [
    "Month", "Platform", "Guest Name", "Check-In", "Check-Out",
    "Nights", "Nightly Rate", "Cleaning Fee", "Gross Revenue",
    "Platform Fees", "Net Payout", "Notes"
]
for c, h in enumerate(bk_headers, 1):
    ws_bk.cell(row=1, column=c, value=h)
style_header(ws_bk, 1, len(bk_headers))

# Restore data
for i, vals in enumerate(existing_bookings):
    r = 2 + i
    for c, v in enumerate(vals, 1):
        if v is not None and not str(v).startswith('='):
            ws_bk.cell(row=r, column=c, value=v)

# Format all rows — NO formulas in data columns where values already exist
for r in range(2, BK_LAST + 1):
    ws_bk.cell(row=r, column=4).number_format = DATE_FMT
    ws_bk.cell(row=r, column=5).number_format = DATE_FMT
    for c in [7, 8, 9, 10, 11]:
        ws_bk.cell(row=r, column=c).number_format = MONEY_FMT

style_area(ws_bk, 1, BK_LAST, len(bk_headers))
auto_width(ws_bk, len(bk_headers))
ws_bk.auto_filter.ref = f"A1:L{BK_LAST}"
ws_bk.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2: Expenses
# ═══════════════════════════════════════════════════════════════════════
ws_ex = wb.create_sheet("Expenses")
ws_ex.sheet_properties.tabColor = "8B0000"

ex_headers = [
    "Date", "Month", "Category", "Vendor/Payee", "Description",
    "Amount", "Payment Method", "Receipt? (Y/N)", "Notes"
]
for c, h in enumerate(ex_headers, 1):
    ws_ex.cell(row=1, column=c, value=h)
style_header(ws_ex, 1, len(ex_headers))

# Dropdowns
cat_dv = DataValidation(type="list", formula1='"' + ','.join(CATEGORIES) + '"', allow_blank=True)
cat_dv.error = "Pick a Schedule E category"
ws_ex.add_data_validation(cat_dv)

pay_dv = DataValidation(type="list", formula1='"Credit Card,Debit Card,Cash,Check,Venmo,Zelle,PayPal,Other"', allow_blank=True)
ws_ex.add_data_validation(pay_dv)

rcpt_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws_ex.add_data_validation(rcpt_dv)

# Restore data
for i, vals in enumerate(existing_expenses):
    r = 2 + i
    for c, v in enumerate(vals, 1):
        if v is not None and not str(v).startswith('='):
            ws_ex.cell(row=r, column=c, value=v)

# Format + Month formula for all rows
for r in range(2, EX_LAST + 1):
    ws_ex.cell(row=r, column=1).number_format = DATE_FMT
    # Month auto-derives from Date — always a formula
    ws_ex.cell(row=r, column=2).value = f'=IF(A{r}="","",TEXT(A{r},"MMM YYYY"))'
    ws_ex.cell(row=r, column=6).number_format = MONEY_FMT
    cat_dv.add(ws_ex.cell(row=r, column=3))
    pay_dv.add(ws_ex.cell(row=r, column=7))
    rcpt_dv.add(ws_ex.cell(row=r, column=8))

style_area(ws_ex, 1, EX_LAST, len(ex_headers))
auto_width(ws_ex, len(ex_headers))
ws_ex.auto_filter.ref = f"A1:I{EX_LAST}"
ws_ex.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 3: Monthly Summary — uses SUMIFS (not SUMPRODUCT)
# ═══════════════════════════════════════════════════════════════════════
ws_ms = wb.create_sheet("Monthly Summary")
ws_ms.sheet_properties.tabColor = "4472C4"

ws_ms.cell(row=1, column=1, value="Keng's Landing — Monthly P&L Summary").font = TITLE_FONT
ws_ms.merge_cells("A1:G1")

ms_headers = ["Month", "Revenue", "Expenses", "Net Profit/Loss", "Bookings", "Nights Booked", "Occupancy Rate"]
for c, h in enumerate(ms_headers, 1):
    ws_ms.cell(row=3, column=c, value=h)
style_header(ws_ms, 3, len(ms_headers))

# Booking ranges as absolute refs
BK_A = f"Bookings!$A$2:$A${BK_LAST}"
BK_K = f"Bookings!$K$2:$K${BK_LAST}"
BK_F = f"Bookings!$F$2:$F${BK_LAST}"
EX_B = f"Expenses!$B$2:$B${EX_LAST}"
EX_F = f"Expenses!$F$2:$F${EX_LAST}"
EX_C = f"Expenses!$C$2:$C${EX_LAST}"

for i, month in enumerate(MONTHS_2026):
    r = 4 + i
    ws_ms.cell(row=r, column=1, value=month)

    # Revenue = sum of Net Payout where Bookings Month matches
    ws_ms.cell(row=r, column=2).value = f'=SUMIFS({BK_K},{BK_A},A{r})'
    ws_ms.cell(row=r, column=2).number_format = MONEY_FMT

    # Expenses = sum of Amount where Expenses Month matches
    ws_ms.cell(row=r, column=3).value = f'=SUMIFS({EX_F},{EX_B},A{r})'
    ws_ms.cell(row=r, column=3).number_format = MONEY_FMT

    # Net Profit/Loss
    ws_ms.cell(row=r, column=4).value = f'=B{r}-C{r}'
    ws_ms.cell(row=r, column=4).number_format = MONEY_FMT

    # Booking count
    ws_ms.cell(row=r, column=5).value = f'=COUNTIFS({BK_A},A{r})'

    # Nights booked
    ws_ms.cell(row=r, column=6).value = f'=SUMIFS({BK_F},{BK_A},A{r})'

    # Occupancy rate
    ws_ms.cell(row=r, column=7).value = f'=IFERROR(F{r}/{DAYS_IN_MONTH[i]},0)'
    ws_ms.cell(row=r, column=7).number_format = PCT_FMT

# Annual totals row
TR = 16
ws_ms.cell(row=TR, column=1, value="TOTAL 2026").font = BOLD
for c in [2, 3, 4]:
    ws_ms.cell(row=TR, column=c).value = f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}15)'
    ws_ms.cell(row=TR, column=c).number_format = MONEY_FMT
    ws_ms.cell(row=TR, column=c).font = BOLD
ws_ms.cell(row=TR, column=5).value = '=SUM(E4:E15)'
ws_ms.cell(row=TR, column=5).font = BOLD
ws_ms.cell(row=TR, column=6).value = '=SUM(F4:F15)'
ws_ms.cell(row=TR, column=6).font = BOLD
ws_ms.cell(row=TR, column=7).value = '=IFERROR(F16/365,0)'
ws_ms.cell(row=TR, column=7).number_format = PCT_FMT
ws_ms.cell(row=TR, column=7).font = BOLD

# ── Expense breakdown by category ──
ws_ms.cell(row=18, column=1, value="Expense Breakdown by Category").font = TITLE_FONT
ws_ms.merge_cells("A18:C18")

for c, h in enumerate(["Category", "YTD Total", "% of Total"], 1):
    ws_ms.cell(row=19, column=c, value=h)
style_header(ws_ms, 19, 3)

for i, cat in enumerate(CATEGORIES):
    r = 20 + i
    ws_ms.cell(row=r, column=1, value=cat)
    ws_ms.cell(row=r, column=2).value = f'=SUMIFS({EX_F},{EX_C},A{r})'
    ws_ms.cell(row=r, column=2).number_format = MONEY_FMT
    ws_ms.cell(row=r, column=3).value = f'=IFERROR(B{r}/C{TR},0)'
    ws_ms.cell(row=r, column=3).number_format = PCT_FMT

ct_row = 20 + len(CATEGORIES)
ws_ms.cell(row=ct_row, column=1, value="Total").font = BOLD
ws_ms.cell(row=ct_row, column=2).value = f'=SUM(B20:B{ct_row - 1})'
ws_ms.cell(row=ct_row, column=2).number_format = MONEY_FMT
ws_ms.cell(row=ct_row, column=2).font = BOLD

style_area(ws_ms, 3, TR, len(ms_headers))
style_area(ws_ms, 19, ct_row, 3)
auto_width(ws_ms, len(ms_headers), minw=16)
ws_ms.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 4: Investment & ROI — all auto from Monthly Summary
# ═══════════════════════════════════════════════════════════════════════
ws_roi = wb.create_sheet("Investment & ROI")
ws_roi.sheet_properties.tabColor = "FFB900"

ws_roi.cell(row=1, column=1, value="Keng's Landing — Investment Recovery Tracker").font = TITLE_FONT
ws_roi.merge_cells("A1:D1")

# Investment
ws_roi.cell(row=3, column=1, value="INITIAL INVESTMENT").font = SUBHEADER_FONT
ws_roi.cell(row=3, column=1).fill = SUBHEADER_FILL
ws_roi.cell(row=3, column=2).fill = SUBHEADER_FILL

ws_roi.cell(row=4, column=1, value="Property Purchase (4 acres + house, hot tub, pond)")
ws_roi.cell(row=4, column=2, value=87000).number_format = MONEY_FMT

ws_roi.cell(row=5, column=1, value="Improvements")
ws_roi.cell(row=5, column=2, value=10000).number_format = MONEY_FMT

ws_roi.cell(row=6, column=1, value="Total Investment").font = BOLD
ws_roi.cell(row=6, column=2).value = '=B4+B5'
ws_roi.cell(row=6, column=2).number_format = MONEY_FMT
ws_roi.cell(row=6, column=2).font = BOLD

# ROI targets
ws_roi.cell(row=8, column=1, value="ROI TARGET").font = SUBHEADER_FONT
ws_roi.cell(row=8, column=1).fill = SUBHEADER_FILL
ws_roi.cell(row=8, column=2).fill = SUBHEADER_FILL

ws_roi.cell(row=9, column=1, value="Target payback (years)")
ws_roi.cell(row=9, column=2, value=7)

ws_roi.cell(row=10, column=1, value="Required annual net profit")
ws_roi.cell(row=10, column=2).value = '=B6/B9'
ws_roi.cell(row=10, column=2).number_format = MONEY_FMT

ws_roi.cell(row=11, column=1, value="Required monthly net profit")
ws_roi.cell(row=11, column=2).value = '=B10/12'
ws_roi.cell(row=11, column=2).number_format = MONEY_FMT

ws_roi.cell(row=12, column=1, value="At $150/night, min nights/month needed")
ws_roi.cell(row=12, column=2).value = '=ROUNDUP(B11/150,0)'

# Progress — auto from Monthly Summary
ws_roi.cell(row=14, column=1, value="PROGRESS (auto-calculated)").font = SUBHEADER_FONT
ws_roi.cell(row=14, column=1).fill = SUBHEADER_FILL
ws_roi.cell(row=14, column=2).fill = SUBHEADER_FILL

ws_roi.cell(row=15, column=1, value="Total revenue to date")
ws_roi.cell(row=15, column=2).value = "='Monthly Summary'!B16"
ws_roi.cell(row=15, column=2).number_format = MONEY_FMT

ws_roi.cell(row=16, column=1, value="Total expenses to date")
ws_roi.cell(row=16, column=2).value = "='Monthly Summary'!C16"
ws_roi.cell(row=16, column=2).number_format = MONEY_FMT

ws_roi.cell(row=17, column=1, value="Net profit to date").font = BOLD
ws_roi.cell(row=17, column=2).value = "='Monthly Summary'!D16"
ws_roi.cell(row=17, column=2).number_format = MONEY_FMT
ws_roi.cell(row=17, column=2).font = BOLD

ws_roi.cell(row=18, column=1, value="Remaining to recover")
ws_roi.cell(row=18, column=2).value = '=B6-B17'
ws_roi.cell(row=18, column=2).number_format = MONEY_FMT

ws_roi.cell(row=19, column=1, value="% recovered").font = BOLD_GREEN
ws_roi.cell(row=19, column=2).value = '=IFERROR(B17/B6,0)'
ws_roi.cell(row=19, column=2).number_format = PCT_FMT
ws_roi.cell(row=19, column=2).font = BOLD_GREEN

ws_roi.cell(row=20, column=1, value="Total bookings to date")
ws_roi.cell(row=20, column=2).value = "='Monthly Summary'!E16"

ws_roi.cell(row=21, column=1, value="Annual occupancy rate")
ws_roi.cell(row=21, column=2).value = "='Monthly Summary'!G16"
ws_roi.cell(row=21, column=2).number_format = PCT_FMT

# Pace projections
ws_roi.cell(row=23, column=1, value="PACE PROJECTIONS").font = SUBHEADER_FONT
ws_roi.cell(row=23, column=1).fill = SUBHEADER_FILL
ws_roi.cell(row=23, column=2).fill = SUBHEADER_FILL

ws_roi.cell(row=24, column=1, value="Months with revenue")
ws_roi.cell(row=24, column=2).value = '=COUNTIF(\'Monthly Summary\'!B4:B15,">0")'

ws_roi.cell(row=25, column=1, value="Average monthly net profit")
ws_roi.cell(row=25, column=2).value = '=IFERROR(B17/B24,0)'
ws_roi.cell(row=25, column=2).number_format = MONEY_FMT

ws_roi.cell(row=26, column=1, value="Projected annual net profit (at current pace)")
ws_roi.cell(row=26, column=2).value = '=B25*12'
ws_roi.cell(row=26, column=2).number_format = MONEY_FMT

ws_roi.cell(row=27, column=1, value="Projected years to recover investment")
ws_roi.cell(row=27, column=2).value = '=IFERROR(B6/B26,0)'
ws_roi.cell(row=27, column=2).number_format = '0.0'

ws_roi.cell(row=28, column=1, value="On track for 7-year target?").font = BOLD
ws_roi.cell(row=28, column=2).value = '=IF(B24=0,"No data yet",IF(B27<=B9,"YES","NO — need more revenue"))'
ws_roi.cell(row=28, column=2).font = BOLD

# Expansion notes
ws_roi.cell(row=30, column=1, value="FUTURE EXPANSION NOTES").font = SUBHEADER_FONT
ws_roi.cell(row=30, column=1).fill = SUBHEADER_FILL
ws_roi.merge_cells("A30:D30")
for c in [2, 3, 4]:
    ws_roi.cell(row=30, column=c).fill = SUBHEADER_FILL

for i, note in enumerate([
    "Tiny homes / mobile homes — additional rental units on the 4 acres",
    "Infrastructure needed: gravel paths, septic, water, electric meter",
    "May be tracked as separate venture/series — TBD with CPA",
    "Julie Rose exploring property management role ($20/hr)",
]):
    ws_roi.cell(row=31 + i, column=1, value=f"• {note}")
    ws_roi.merge_cells(f"A{31+i}:D{31+i}")

ws_roi.column_dimensions['A'].width = 50
ws_roi.column_dimensions['B'].width = 22

# ═══════════════════════════════════════════════════════════════════════
# SHEET 5: Budget vs Actual
# ═══════════════════════════════════════════════════════════════════════
ws_bgt = wb.create_sheet("Budget vs Actual")
ws_bgt.sheet_properties.tabColor = "7030A0"  # purple

ws_bgt.cell(row=1, column=1, value="Keng's Landing — Annual Budget vs Actual").font = TITLE_FONT
ws_bgt.merge_cells("A1:F1")

ws_bgt.cell(row=2, column=1, value="Set your Annual Budget per category in column B. Everything else auto-calculates.").font = Font(italic=True, color="666666")
ws_bgt.merge_cells("A2:F2")

bgt_headers = ["Category", "Annual Budget", "Monthly Budget", "YTD Actual", "YTD Variance", "% Used"]
for c, h in enumerate(bgt_headers, 1):
    ws_bgt.cell(row=4, column=c, value=h)
style_header(ws_bgt, 4, len(bgt_headers))

# Same expense ranges as Monthly Summary
for i, cat in enumerate(CATEGORIES):
    r = 5 + i
    ws_bgt.cell(row=r, column=1, value=cat)

    # B: Annual Budget — user enters this (restore existing if available)
    if cat in existing_budgets:
        ws_bgt.cell(row=r, column=2, value=existing_budgets[cat])
    ws_bgt.cell(row=r, column=2).number_format = MONEY_FMT

    # C: Monthly Budget = Annual / 12
    ws_bgt.cell(row=r, column=3).value = f'=IFERROR(B{r}/12,0)'
    ws_bgt.cell(row=r, column=3).number_format = MONEY_FMT

    # D: YTD Actual — SUMIFS from Expenses (same formula as Monthly Summary category breakdown)
    ws_bgt.cell(row=r, column=4).value = f'=SUMIFS({EX_F},{EX_C},A{r})'
    ws_bgt.cell(row=r, column=4).number_format = MONEY_FMT

    # E: Variance = Budget - Actual (positive = under budget, negative = over)
    ws_bgt.cell(row=r, column=5).value = f'=B{r}-D{r}'
    ws_bgt.cell(row=r, column=5).number_format = MONEY_FMT

    # F: % Used = Actual / Budget
    ws_bgt.cell(row=r, column=6).value = f'=IFERROR(D{r}/B{r},0)'
    ws_bgt.cell(row=r, column=6).number_format = PCT_FMT

# Totals row
bgt_total = 5 + len(CATEGORIES)
ws_bgt.cell(row=bgt_total, column=1, value="TOTAL").font = BOLD
for c in [2, 3, 4, 5]:
    ws_bgt.cell(row=bgt_total, column=c).value = f'=SUM({get_column_letter(c)}5:{get_column_letter(c)}{bgt_total - 1})'
    ws_bgt.cell(row=bgt_total, column=c).number_format = MONEY_FMT
    ws_bgt.cell(row=bgt_total, column=c).font = BOLD
ws_bgt.cell(row=bgt_total, column=6).value = f'=IFERROR(D{bgt_total}/B{bgt_total},0)'
ws_bgt.cell(row=bgt_total, column=6).number_format = PCT_FMT
ws_bgt.cell(row=bgt_total, column=6).font = BOLD

# Revenue budget section
rev_start = bgt_total + 2
ws_bgt.cell(row=rev_start, column=1, value="Revenue Target").font = SUBHEADER_FONT
ws_bgt.cell(row=rev_start, column=1).fill = SUBHEADER_FILL
ws_bgt.merge_cells(f"A{rev_start}:F{rev_start}")
for c in range(2, 7):
    ws_bgt.cell(row=rev_start, column=c).fill = SUBHEADER_FILL

rev_headers = ["Metric", "Annual Target", "Monthly Target", "YTD Actual", "YTD Variance", "% Achieved"]
for c, h in enumerate(rev_headers, 1):
    ws_bgt.cell(row=rev_start + 1, column=c, value=h)
style_header(ws_bgt, rev_start + 1, len(rev_headers))

rr = rev_start + 2
ws_bgt.cell(row=rr, column=1, value="Gross Revenue")
ws_bgt.cell(row=rr, column=2).number_format = MONEY_FMT  # user sets annual target
ws_bgt.cell(row=rr, column=3).value = f'=IFERROR(B{rr}/12,0)'
ws_bgt.cell(row=rr, column=3).number_format = MONEY_FMT
ws_bgt.cell(row=rr, column=4).value = "='Monthly Summary'!B16"
ws_bgt.cell(row=rr, column=4).number_format = MONEY_FMT
ws_bgt.cell(row=rr, column=5).value = f'=D{rr}-B{rr}'
ws_bgt.cell(row=rr, column=5).number_format = MONEY_FMT
ws_bgt.cell(row=rr, column=6).value = f'=IFERROR(D{rr}/B{rr},0)'
ws_bgt.cell(row=rr, column=6).number_format = PCT_FMT

rr2 = rr + 1
ws_bgt.cell(row=rr2, column=1, value="Net Profit")
ws_bgt.cell(row=rr2, column=2).number_format = MONEY_FMT  # user sets annual target
ws_bgt.cell(row=rr2, column=3).value = f'=IFERROR(B{rr2}/12,0)'
ws_bgt.cell(row=rr2, column=3).number_format = MONEY_FMT
ws_bgt.cell(row=rr2, column=4).value = "='Monthly Summary'!D16"
ws_bgt.cell(row=rr2, column=4).number_format = MONEY_FMT
ws_bgt.cell(row=rr2, column=5).value = f'=D{rr2}-B{rr2}'
ws_bgt.cell(row=rr2, column=5).number_format = MONEY_FMT
ws_bgt.cell(row=rr2, column=6).value = f'=IFERROR(D{rr2}/B{rr2},0)'
ws_bgt.cell(row=rr2, column=6).number_format = PCT_FMT

style_area(ws_bgt, 4, bgt_total, len(bgt_headers))
style_area(ws_bgt, rev_start + 1, rr2, len(rev_headers))

ws_bgt.column_dimensions['A'].width = 24
for c in range(2, 7):
    ws_bgt.column_dimensions[get_column_letter(c)].width = 16
ws_bgt.freeze_panes = "A5"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 6: Category Reference
# ═══════════════════════════════════════════════════════════════════════
ws_ref = wb.create_sheet("Category Reference")
ws_ref.sheet_properties.tabColor = "808080"

ws_ref.cell(row=1, column=1, value="Schedule E Expense Categories — Quick Reference").font = TITLE_FONT
ws_ref.merge_cells("A1:C1")

for c, h in enumerate(["Category", "What Counts", "Examples"], 1):
    ws_ref.cell(row=3, column=c, value=h)
style_header(ws_ref, 3, 3)

for i, (cat, what, ex) in enumerate([
    ("Mortgage Interest", "Interest portion of monthly mortgage only", "Bank statement line item"),
    ("Property Taxes", "County/city property taxes", "Freestone County tax bill"),
    ("Insurance", "Property insurance premiums", "Landlord policy, umbrella policy"),
    ("Repairs & Maintenance", "Fixing or maintaining existing property", "Plumbing, HVAC repair, pest control, hot tub maintenance"),
    ("Supplies", "Consumables for guests or property", "Cleaning supplies, toiletries, linens, kitchen items, light bulbs"),
    ("Utilities", "Monthly service bills", "Electric, water, propane, Starlink internet, trash"),
    ("Cleaning & Turnover", "Cleaning between guests", "Professional cleaner per-visit fee (Julie Rose)"),
    ("Platform Fees", "Host fees charged by platforms", "Airbnb 3% host fee (if separate from payout)"),
    ("Professional Services", "Expert help", "CPA, attorney, property manager"),
    ("Advertising", "Marketing spend", "Listing boost, professional photos"),
    ("Travel / Mileage", "Trips to property for business", "IRS standard mileage rate x miles driven"),
    ("Depreciation", "Asset value depreciation", "Building: 27.5yr, Furniture: 5-7yr"),
    ("Other", "Anything that doesn't fit above", "Smart locks, permits, software subscriptions"),
]):
    r = 4 + i
    ws_ref.cell(row=r, column=1, value=cat)
    ws_ref.cell(row=r, column=2, value=what)
    ws_ref.cell(row=r, column=3, value=ex)

style_area(ws_ref, 3, 16, 3)
ws_ref.column_dimensions['A'].width = 22
ws_ref.column_dimensions['B'].width = 40
ws_ref.column_dimensions['C'].width = 50

# ═══════════════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════════════
wb.save(OUTPUT_PATH)
print(f"Created: {OUTPUT_PATH}")
