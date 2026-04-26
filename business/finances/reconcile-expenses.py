#!/usr/bin/env python3
"""
Reconcile reviewed expenses in the finance tracker.

After you've reviewed the yellow "Review" rows in Excel and changed their
Status to either "Business" or "Personal", run this script to:
1. Remove all "Personal" rows from the Expenses sheet
2. Clear the yellow highlighting from reviewed rows
3. Report what was removed

Usage:
    python3 finances/reconcile-expenses.py
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill

TRACKER_PATH = os.path.join(os.path.dirname(__file__), "kengs-landing-finance-tracker.xlsx")
NO_FILL = PatternFill(fill_type=None)


def main():
    if not os.path.exists(TRACKER_PATH):
        print(f"Error: Tracker not found: {TRACKER_PATH}")
        return

    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb["Expenses"]

    # Scan all rows, collect personal ones to remove
    personal_rows = []
    still_review = 0
    business_count = 0

    for r in range(2, 202):
        if ws.cell(row=r, column=1).value is None:
            break
        status = ws.cell(row=r, column=9).value or ""

        if status == "Personal":
            date = ws.cell(row=r, column=1).value
            vendor = ws.cell(row=r, column=4).value
            amt = ws.cell(row=r, column=6).value
            personal_rows.append({"row": r, "date": str(date)[:10], "vendor": vendor, "amount": amt})
        elif status == "Review":
            still_review += 1
        elif status == "Business":
            business_count += 1
            # Clear yellow highlight if it was previously a Review item
            for c in range(1, 10):
                ws.cell(row=r, column=c).fill = NO_FILL

    if not personal_rows:
        print("No items marked 'Personal' found.")
        if still_review > 0:
            print(f"  {still_review} items still marked 'Review' — update them in Excel first.")
        else:
            print("  All items are marked 'Business'. Nothing to do.")
        return

    # Report what will be removed
    print(f"Removing {len(personal_rows)} personal expense(s):")
    total_removed = 0
    for p in personal_rows:
        print(f"  Row {p['row']}: {p['date']}  {str(p['vendor']):25s}  ${p['amount']:>8.2f}")
        total_removed += p["amount"]
    print(f"  Total removed: ${total_removed:,.2f}")

    # Delete rows from bottom to top so indices stay valid
    for p in sorted(personal_rows, key=lambda x: x["row"], reverse=True):
        ws.delete_rows(p["row"])

    wb.save(TRACKER_PATH)

    print(f"\nDone! Tracker updated.")
    print(f"  Remaining business expenses: {business_count}")
    if still_review > 0:
        print(f"  Still pending review: {still_review}")
    print(f"  Tracker: {TRACKER_PATH}")


if __name__ == "__main__":
    main()
