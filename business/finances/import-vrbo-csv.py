#!/usr/bin/env python3
"""
Import VRBO/HomeAway CSV export into the Keng's Landing finance tracker.

Usage:
    python3 finances/import-vrbo-csv.py [path-to-vrbo-csv]
    python3 finances/import-vrbo-csv.py --detect [path-to-vrbo-csv]

If no path is given, auto-detects the newest vrbo*.csv or homeaway*.csv in ~/Downloads/.

Use --detect to print the CSV column headers found in the file without importing.
This helps calibrate the column mappings if VRBO changes their export format.

VRBO CSV column names vary slightly by export type and region. The script uses
a flexible alias map to handle known variations. If a required column is not
found, the script prints a clear error and lists what columns were detected.

Writes new/enriched entries to the Bookings sheet in the finance tracker.
Deduplicates by:
  1. Confirmation code (written into the Notes column on previous imports)
  2. Net payout amount + check-in month (for stub entries with no confirmation code)
"""

import csv
import glob
import sys
import os
from datetime import datetime

import openpyxl

# --- Configuration ---
TRACKER_PATH = os.path.join(os.path.dirname(__file__), "kengs-landing-finance-tracker.xlsx")
DOWNLOADS_DIR = os.path.expanduser("~/Downloads")
SHEET_NAME = "Bookings"
MAX_ROWS = 101  # Header + 100 data rows

# Bookings column indices (1-based) — must match the tracker sheet layout
COL_MONTH = 1       # A: "MMM YYYY"
COL_PLATFORM = 2    # B: "VRBO"
COL_GUEST = 3       # C: Guest Name
COL_CHECKIN = 4     # D: Check-In date
COL_CHECKOUT = 5    # E: Check-Out date
COL_NIGHTS = 6      # F: Nights
COL_RATE = 7        # G: Nightly Rate (calculated)
COL_CLEANING = 8    # H: Cleaning Fee
COL_GROSS = 9       # I: Gross Revenue (total charged to guest)
COL_FEES = 10       # J: Platform Fees (VRBO commission/service fee)
COL_NET = 11        # K: Net Payout (what hits the bank)
COL_NOTES = 12      # L: Notes (includes confirmation code)

# ---------------------------------------------------------------------------
# Column alias map
# Keys are our canonical field names. Values are lists of possible VRBO
# column header strings, tried in order (case-insensitive match).
# Run with --detect on a real CSV to see what columns VRBO actually exports,
# then update this map if needed.
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "confirmation_code": [
        "Confirmation Number",
        "Reservation ID",
        "Booking ID",
        "Confirmation Code",
        "ID",
    ],
    "guest_name": [
        "Guest Name",
        "Guest",
        "Traveler Name",
        "Traveler",
        "Renter Name",
    ],
    "checkin": [
        "Check-in",
        "Check In",
        "Arrival",
        "Arrival Date",
        "Check-In Date",
        "Start Date",
    ],
    "checkout": [
        "Check-out",
        "Check Out",
        "Departure",
        "Departure Date",
        "Check-Out Date",
        "End Date",
    ],
    "nights": [
        "Nights",
        "Number of Nights",
        "Duration",
        "Stay Length",
    ],
    "gross": [
        "Traveler Payments",
        "Total Traveler Payment",
        "Gross Revenue",
        "Total Amount",
        "Total Charged",
        "Total",
        "Rent",
        "Base Rent",
    ],
    "cleaning_fee": [
        "Cleaning Fee",
        "Cleaning",
    ],
    "pet_fee": [
        "Pet Fee",
        "Pet",
    ],
    "platform_fee": [
        "Channel Commission",
        "Host Service Fee",
        "Commission",
        "Service Fee",
        "VRBO Commission",
        "Vrbo Commission",
        "Platform Fee",
        "Owner Fees",
    ],
    "net_payout": [
        "Owner Revenue",
        "Net Payout",
        "Net Revenue",
        "Payout",
        "Owner Payout",
        "Amount Paid to Owner",
        "Your Earnings",
    ],
    "status": [
        "Status",
        "Reservation Status",
        "Booking Status",
    ],
    "booking_date": [
        "Booking Date",
        "Date Booked",
        "Created",
        "Reservation Date",
    ],
}

# Statuses that represent a real paid booking (ignore cancellations, inquiries)
VALID_STATUSES = {
    "confirmed",
    "completed",
    "checked out",
    "booked",
    "accepted",
    "paid",
    "reserved",
}


# ---------------------------------------------------------------------------
# Column mapping helpers
# ---------------------------------------------------------------------------

def build_column_map(headers):
    """Map our canonical field names to actual CSV column names.

    Returns:
        col_map: dict of canonical_name -> actual_header_string (or None if not found)
        missing_required: list of canonical names for required fields not found
    """
    headers_lower = {h.strip().lower(): h.strip() for h in headers}
    col_map = {}
    required = {"confirmation_code", "guest_name", "checkin", "checkout"}
    missing_required = []

    for field, aliases in COLUMN_ALIASES.items():
        matched = None
        for alias in aliases:
            if alias.strip().lower() in headers_lower:
                matched = headers_lower[alias.strip().lower()]
                break
        col_map[field] = matched
        if matched is None and field in required:
            missing_required.append(field)

    return col_map, missing_required


def get_val(row, col_map, field):
    """Get a value from a CSV row dict using the canonical field name."""
    col = col_map.get(field)
    if col is None:
        return ""
    return row.get(col, "").strip()


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def parse_date(date_str):
    """Parse common date formats from VRBO CSVs."""
    if not date_str or not date_str.strip():
        return None
    date_str = date_str.strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def to_month_str(dt):
    """Convert datetime to 'MMM YYYY' format matching the tracker."""
    return dt.strftime("%b %Y")


def parse_float(val):
    """Parse a currency/numeric string, returning 0.0 for empty/None."""
    if not val or val.strip() == "":
        return 0.0
    # Strip currency symbols and commas
    cleaned = val.strip().lstrip("$").replace(",", "").replace("(", "-").replace(")", "")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# CSV reading
# ---------------------------------------------------------------------------

def read_vrbo_reservations(csv_path, col_map):
    """Read VRBO CSV and return valid reservation rows as dicts.

    Filters out cancellations and non-reservation rows by checking the
    Status column if present.
    """
    reservations = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Skip rows where status indicates cancellation/inquiry
            status_col = col_map.get("status")
            if status_col:
                status = row.get(status_col, "").strip().lower()
                if status and status not in VALID_STATUSES:
                    continue

            # Skip rows with no check-in date (summary/footer rows)
            checkin_val = get_val(row, col_map, "checkin")
            if not checkin_val:
                continue

            reservations.append(row)
    return reservations


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def find_existing_bookings(ws):
    """Scan Bookings sheet and return existing entries for dedup.

    Returns:
        confirmation_codes: set of confirmation codes in Notes column
        payout_month_map: dict of (net_payout_float, month_str) -> row_number
            for stub entries with no guest name
        next_empty_row: first row with no data in column A
    """
    confirmation_codes = set()
    payout_month_map = {}
    next_empty_row = None

    for r in range(2, MAX_ROWS + 1):
        month_val = ws.cell(row=r, column=COL_MONTH).value
        if month_val is None:
            if next_empty_row is None:
                next_empty_row = r
            continue

        notes = ws.cell(row=r, column=COL_NOTES).value or ""
        net_val = ws.cell(row=r, column=COL_NET).value

        # Extract confirmation codes from notes.
        # VRBO IDs often start with "HA" or are purely numeric.
        for word in notes.split():
            if (word.startswith("HA") and len(word) >= 6) or word.isdigit():
                confirmation_codes.add(word)
            # Also handle the "VRBO XXXXXXXX" prefix format
            if word.startswith("VRBO") and len(word) > 4:
                confirmation_codes.add(word[4:].lstrip("-").strip())

        # Track stub entries (same net payout + month, no guest name)
        guest = ws.cell(row=r, column=COL_GUEST).value
        if net_val and not guest:
            try:
                key = (round(float(net_val), 2), str(month_val).strip())
                payout_month_map[key] = r
            except (ValueError, TypeError):
                pass

    if next_empty_row is None:
        next_empty_row = MAX_ROWS + 1

    return confirmation_codes, payout_month_map, next_empty_row


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def write_booking(ws, row_num, row, col_map):
    """Write a single VRBO reservation to a row in the Bookings sheet."""
    conf_code = get_val(row, col_map, "confirmation_code")
    guest = get_val(row, col_map, "guest_name")
    checkin = parse_date(get_val(row, col_map, "checkin"))
    checkout = parse_date(get_val(row, col_map, "checkout"))

    nights_str = get_val(row, col_map, "nights")
    nights = int(float(nights_str)) if nights_str else None
    if nights is None and checkin and checkout:
        nights = (checkout - checkin).days

    gross = parse_float(get_val(row, col_map, "gross"))
    cleaning_fee = parse_float(get_val(row, col_map, "cleaning_fee"))
    pet_fee = parse_float(get_val(row, col_map, "pet_fee"))
    platform_fee = parse_float(get_val(row, col_map, "platform_fee"))
    net_payout = parse_float(get_val(row, col_map, "net_payout"))

    # Fall back: derive net payout if not directly available
    if net_payout == 0.0 and gross > 0:
        net_payout = gross - platform_fee

    # Nightly rate = (Gross - Cleaning - Pet) / Nights
    nightly_rate = None
    if nights and nights > 0 and gross > 0:
        nightly_rate = round((gross - cleaning_fee - pet_fee) / nights, 2)

    month_str = to_month_str(checkin) if checkin else ""

    ws.cell(row=row_num, column=COL_MONTH, value=month_str)
    ws.cell(row=row_num, column=COL_PLATFORM, value="VRBO")
    ws.cell(row=row_num, column=COL_GUEST, value=guest)

    if checkin:
        ws.cell(row=row_num, column=COL_CHECKIN, value=checkin)
        ws.cell(row=row_num, column=COL_CHECKIN).number_format = "YYYY-MM-DD"
    if checkout:
        ws.cell(row=row_num, column=COL_CHECKOUT, value=checkout)
        ws.cell(row=row_num, column=COL_CHECKOUT).number_format = "YYYY-MM-DD"

    if nights:
        ws.cell(row=row_num, column=COL_NIGHTS, value=nights)
    if nightly_rate:
        ws.cell(row=row_num, column=COL_RATE, value=nightly_rate)
        ws.cell(row=row_num, column=COL_RATE).number_format = "#,##0.00"
    if cleaning_fee:
        ws.cell(row=row_num, column=COL_CLEANING, value=cleaning_fee)
        ws.cell(row=row_num, column=COL_CLEANING).number_format = "#,##0.00"

    ws.cell(row=row_num, column=COL_GROSS, value=gross)
    ws.cell(row=row_num, column=COL_GROSS).number_format = "#,##0.00"
    ws.cell(row=row_num, column=COL_FEES, value=platform_fee)
    ws.cell(row=row_num, column=COL_FEES).number_format = "#,##0.00"
    ws.cell(row=row_num, column=COL_NET, value=net_payout)
    ws.cell(row=row_num, column=COL_NET).number_format = "#,##0.00"

    note_parts = [f"VRBO {conf_code}"]
    if pet_fee > 0:
        note_parts.append(f"Pet fee: ${pet_fee:.2f}")
    ws.cell(row=row_num, column=COL_NOTES, value=" | ".join(note_parts))


# ---------------------------------------------------------------------------
# Auto-detect CSV
# ---------------------------------------------------------------------------

def find_latest_vrbo_csv():
    """Find the newest vrbo*.csv or homeaway*.csv in ~/Downloads/."""
    patterns = [
        os.path.join(DOWNLOADS_DIR, "vrbo*.csv"),
        os.path.join(DOWNLOADS_DIR, "VRBO*.csv"),
        os.path.join(DOWNLOADS_DIR, "Vrbo*.csv"),
        os.path.join(DOWNLOADS_DIR, "homeaway*.csv"),
        os.path.join(DOWNLOADS_DIR, "HomeAway*.csv"),
        os.path.join(DOWNLOADS_DIR, "reservations*.csv"),
    ]
    matches = []
    for pattern in patterns:
        matches.extend(glob.glob(pattern, recursive=False))
    if not matches:
        return None
    matches.sort(key=os.path.getmtime, reverse=True)
    return matches[0]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    detect_only = "--detect" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]

    if args:
        csv_path = args[0]
    else:
        csv_path = find_latest_vrbo_csv()
        if csv_path:
            print(f"Auto-detected: {csv_path}")
            print(f"  (modified: {datetime.fromtimestamp(os.path.getmtime(csv_path)).strftime('%Y-%m-%d %H:%M')})")
        else:
            print("No vrbo*.csv or homeaway*.csv found in ~/Downloads/.")
            print("Usage: python3 finances/import-vrbo-csv.py [path-to-vrbo-csv]")
            print("       python3 finances/import-vrbo-csv.py --detect [path-to-vrbo-csv]")
            sys.exit(1)

    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found: {csv_path}")
        sys.exit(1)

    # Peek at headers first
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []

    if detect_only:
        print(f"\nColumns detected in {os.path.basename(csv_path)}:")
        for i, h in enumerate(headers, 1):
            print(f"  {i:2}. {h}")
        print(f"\nTotal: {len(headers)} columns")
        print("\nColumn mapping resolution:")
        col_map, missing = build_column_map(headers)
        for field, resolved in col_map.items():
            status = f"-> {resolved}" if resolved else "NOT FOUND"
            print(f"  {field:20} {status}")
        if missing:
            print(f"\nRequired fields not found: {missing}")
            print("Update COLUMN_ALIASES in the script to map these fields.")
        return

    col_map, missing_required = build_column_map(headers)
    if missing_required:
        print(f"Error: Required columns not found in CSV: {missing_required}")
        print("\nDetected columns:")
        for h in headers:
            print(f"  {h}")
        print("\nRun with --detect flag for full column mapping diagnostics.")
        print("Update COLUMN_ALIASES in the script to match the actual column names.")
        sys.exit(1)

    if not os.path.exists(TRACKER_PATH):
        print(f"Error: Finance tracker not found: {TRACKER_PATH}")
        sys.exit(1)

    # --- Read CSV ---
    reservations = read_vrbo_reservations(csv_path, col_map)
    print(f"Found {len(reservations)} reservation(s) in CSV")

    if not reservations:
        print("No reservations to import.")
        print("(Check that status values match VALID_STATUSES, or that the CSV has check-in dates)")
        return

    # --- Open tracker ---
    wb = openpyxl.load_workbook(TRACKER_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: Sheet '{SHEET_NAME}' not found in tracker.")
        print(f"Available sheets: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    existing_codes, stub_map, next_row = find_existing_bookings(ws)
    print(f"Existing tracker: {len(existing_codes)} confirmation code(s), {len(stub_map)} stub(s), next empty row: {next_row}")

    added = 0
    enriched = 0
    skipped = 0

    for res in reservations:
        conf_code = get_val(res, col_map, "confirmation_code")
        checkin = parse_date(get_val(res, col_map, "checkin"))
        month_str = to_month_str(checkin) if checkin else ""
        net_raw = get_val(res, col_map, "net_payout")
        net_amount = parse_float(net_raw) if net_raw else 0.0

        guest_label = get_val(res, col_map, "guest_name") or "(unknown guest)"

        # Dedup 1: Confirmation code already in tracker notes
        if conf_code in existing_codes:
            print(f"  SKIP (already imported): {conf_code} — {guest_label}")
            skipped += 1
            continue

        # Dedup 2: Stub entry (same net payout + month, no guest name)
        stub_key = (round(net_amount, 2), month_str)
        stub_row = stub_map.get(stub_key)
        if stub_row:
            print(f"  ENRICH row {stub_row}: {conf_code} — {guest_label} (${net_amount:.2f})")
            write_booking(ws, stub_row, res, col_map)
            del stub_map[stub_key]
            enriched += 1
            continue

        # New entry
        if next_row > MAX_ROWS:
            print(f"  ERROR: Bookings sheet is full (max {MAX_ROWS - 1} rows). Cannot add more.")
            break

        print(f"  ADD row {next_row}: {conf_code} — {guest_label} (${net_amount:.2f})")
        write_booking(ws, next_row, res, col_map)
        next_row += 1
        added += 1

    # --- Save ---
    wb.save(TRACKER_PATH)
    print(f"\nDone! Added: {added}, Enriched: {enriched}, Skipped: {skipped}")
    print(f"Tracker saved: {TRACKER_PATH}")


if __name__ == "__main__":
    main()
