#!/usr/bin/env python3
"""
Reconstruct mileage log from CC transaction data.

Detects trips to the Fairfield/Corsicana area (near 360 County Road)
by finding purchase dates at merchants in those locations.

Usage:
    python3 finances/generate-mileage-log.py [cc-csv-path]

If no path is given, uses the Robinhood CC CSV in Downloads.
"""

import csv
import os
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl

TRACKER_PATH = os.path.join(os.path.dirname(__file__), "kengs-landing-finance-tracker.xlsx")
DOWNLOADS_DIR = os.path.expanduser("~/Downloads")
DEFAULT_CSV = os.path.join(DOWNLOADS_DIR, "684d9305-02c2-4cb6-b639-1cfcad1c6ea2.csv")

STR_LOCATIONS = ["FAIRFIELD", "CORSICANA"]
ROUND_TRIP_MILES = 120
IRS_RATE_2026 = 0.70
HOME_BASE = "Mansfield, TX"
DESTINATION = "360 County Road, Fairfield, TX"


def detect_trips(csv_path):
    trips = defaultdict(list)
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["Type"] not in ("Purchase", "Refund"):
                continue
            desc = row.get("Description", "") or ""
            if any(loc in desc.upper() for loc in STR_LOCATIONS):
                date = row["Date"]
                merchant = row["Merchant"]
                trips[date].append(merchant)
    return trips


def determine_purpose(merchants):
    purposes = set()
    for m in merchants:
        ml = m.lower()
        if any(hw in ml for hw in ["home depot", "ace", "capps", "brenco", "lowe"]):
            purposes.add("maintenance/repairs")
        elif any(sup in ml for sup in ["brookshire", "cooper"]):
            purposes.add("supplies")
        elif any(gas in ml for gas in ["shell", "gulf", "love's", "exxon", "chevron", "conoco", "allsup"]):
            purposes.add("transit fuel")
        else:
            purposes.add("property work")
    return ", ".join(sorted(purposes))


def write_mileage_sheet(trips, tracker_path):
    """Add or update a Mileage Log sheet in the tracker."""
    wb = openpyxl.load_workbook(tracker_path)

    sheet_name = "Mileage Log"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # Header
    headers = ["Date", "Origin", "Destination", "Round Trip Miles",
               "Purpose", "Stops/Evidence", "IRS Rate", "Deduction"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    sorted_dates = sorted(trips.keys())
    row = 2
    total_miles = 0
    total_deduction = 0.0

    for date in sorted_dates:
        stops = trips[date]
        merchants = ", ".join(set(stops))[:60]
        purpose = determine_purpose(stops)
        deduction = ROUND_TRIP_MILES * IRS_RATE_2026

        dt = datetime.strptime(date, "%Y-%m-%d")
        ws.cell(row=row, column=1, value=dt)
        ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=row, column=2, value=HOME_BASE)
        ws.cell(row=row, column=3, value=DESTINATION)
        ws.cell(row=row, column=4, value=ROUND_TRIP_MILES)
        ws.cell(row=row, column=5, value=purpose)
        ws.cell(row=row, column=6, value=merchants)
        ws.cell(row=row, column=7, value=IRS_RATE_2026)
        ws.cell(row=row, column=7).number_format = "$#,##0.00"
        ws.cell(row=row, column=8, value=deduction)
        ws.cell(row=row, column=8).number_format = "$#,##0.00"

        total_miles += ROUND_TRIP_MILES
        total_deduction += deduction
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=row, column=4, value=total_miles)
    ws.cell(row=row, column=4).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=row, column=8, value=total_deduction)
    ws.cell(row=row, column=8).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=row, column=8).number_format = "$#,##0.00"

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(tracker_path)
    return len(sorted_dates), total_miles, total_deduction


def main():
    csv_path = DEFAULT_CSV
    if len(sys.argv) >= 2:
        csv_path = sys.argv[1]

    if not os.path.exists(csv_path):
        print(f"Error: CSV not found: {csv_path}")
        sys.exit(1)

    print(f"Reading: {csv_path}")
    trips = detect_trips(csv_path)
    print(f"Detected {len(trips)} trip dates to Fairfield/Corsicana area")

    if not trips:
        print("No trips found.")
        return

    trip_count, total_miles, total_deduction = write_mileage_sheet(trips, TRACKER_PATH)
    print(f"\nMileage Log sheet written to tracker:")
    print(f"  Trips: {trip_count}")
    print(f"  Total miles: {total_miles}")
    print(f"  Total deduction: ${total_deduction:,.2f}")
    print(f"  IRS rate: ${IRS_RATE_2026}/mile")
    print(f"\nTracker: {TRACKER_PATH}")


if __name__ == "__main__":
    main()
