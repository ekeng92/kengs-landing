#!/usr/bin/env python3
"""
Migrate data from kengs-landing-finance-tracker.xlsx into Supabase.

Reads the Bookings and Expenses sheets, maps columns to the DB schema,
and inserts via the Supabase PostgREST API using the service role key.

Usage:
    python3 backend/db/seed-from-excel.py [--dry-run]

Author: AEON Dev | Created: 2026-04-21
"""

import json
import os
import re
import sys
from datetime import datetime

import openpyxl
import requests

# ── Config ──────────────────────────────────────────────────────────────
SUPABASE_URL = "https://ubfvhzepyizfjmghkhyh.supabase.co"
SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "").strip()
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "../../finances/kengs-landing-finance-tracker.xlsx")

WORKSPACE_ID = "b0604861-b7ae-4f1e-a7cb-fe066d57c623"
PROPERTY_360CR = "0e8ab13c-7976-4b9d-a6c6-3561f7a73f40"

DRY_RUN = "--dry-run" in sys.argv

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

# ── Helpers ─────────────────────────────────────────────────────────────

def fmt_date(val):
    """Convert datetime or string to YYYY-MM-DD, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        # Try to parse common formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return None


def safe_float(val):
    """Convert to float, or None."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """Convert to int, or None."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def extract_confirmation_code(notes):
    """Extract Airbnb confirmation code from notes like 'Airbnb HMTK9FNMQK | ...'"""
    if not notes:
        return None
    match = re.search(r"Airbnb\s+(HM[A-Z0-9]+)", str(notes))
    return match.group(1) if match else None


def extract_tax_amount(notes):
    """Extract tax remitted amount from notes like 'Tax remitted: $40.05'"""
    if not notes:
        return None
    match = re.search(r"Tax remitted:\s*\$?([\d,.]+)", str(notes))
    return float(match.group(1).replace(",", "")) if match else None


def supabase_insert(table, rows):
    """Insert rows into a Supabase table via PostgREST."""
    if not rows:
        print(f"  {table}: no rows to insert")
        return []

    if DRY_RUN:
        print(f"  {table}: would insert {len(rows)} rows (dry run)")
        for i, row in enumerate(rows[:3]):
            print(f"    sample[{i}]: {json.dumps(row, default=str)[:200]}")
        if len(rows) > 3:
            print(f"    ... and {len(rows) - 3} more")
        return rows

    url = f"{SUPABASE_URL}/rest/v1/{table}"
    # Insert in batches of 50 to avoid payload limits
    inserted = []
    batch_size = 50
    for i in range(0, len(rows), batch_size):
        batch = rows[i : i + batch_size]
        resp = requests.post(url, headers=HEADERS, json=batch)
        if resp.status_code in (200, 201):
            inserted.extend(resp.json())
            print(f"  {table}: inserted batch {i // batch_size + 1} ({len(batch)} rows)")
        else:
            print(f"  ERROR {table} batch {i // batch_size + 1}: {resp.status_code} — {resp.text[:300]}")
            sys.exit(1)
    return inserted


# ── Parse Expenses ──────────────────────────────────────────────────────

def parse_expenses(ws):
    """Parse the Expenses sheet into DB-ready dicts."""
    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        date_val = excel_row[0]
        if date_val is None:
            continue  # skip empty rows

        date_str = fmt_date(date_val)
        if not date_str:
            print(f"  WARN: skipping expense with unparseable date: {date_val}")
            continue

        category = excel_row[2]
        merchant = excel_row[3]
        description = excel_row[4]
        amount = safe_float(excel_row[5])
        payment_method = excel_row[6]
        receipt_status = excel_row[7]  # Y/N/CC
        review_state = excel_row[8]    # Business/Personal/Review
        tax_period = excel_row[9]      # Pre-Service/Operational

        if amount is None:
            print(f"  WARN: skipping expense with no amount: {date_str} {merchant}")
            continue

        # Map review_state: anything not Business/Personal becomes Review
        if review_state not in ("Business", "Personal"):
            review_state = "Review"

        # Map tax_period
        if tax_period not in ("Pre-Service", "Operational"):
            tax_period = None

        # Map documentation_status: CC/Y/N
        doc_status = None
        if receipt_status in ("CC", "Y", "N"):
            doc_status = receipt_status

        # Committed if classified, draft if still in review
        status = "committed" if review_state in ("Business", "Personal") else "draft"

        rows.append({
            "workspace_id": WORKSPACE_ID,
            "property_id": PROPERTY_360CR,
            "transaction_date": date_str,
            "merchant_name": str(merchant).strip() if merchant else None,
            "description": str(description).strip() if description else None,
            "category": str(category).strip() if category else None,
            "amount": amount,
            "payment_method": str(payment_method).strip() if payment_method else None,
            "review_state": review_state,
            "tax_period": tax_period,
            "documentation_status": doc_status,
            "needs_receipt_followup": doc_status == "N",
            "status": status,
        })

    return rows


# ── Parse Bookings ──────────────────────────────────────────────────────

def parse_bookings(ws):
    """Parse the Bookings sheet into DB-ready dicts."""
    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        # Skip fully empty rows
        if all(v is None for v in excel_row):
            continue

        month = excel_row[0]
        platform = excel_row[1]
        guest_name = excel_row[2]
        check_in = excel_row[3]
        check_out = excel_row[4]
        nights = safe_int(excel_row[5])
        # nightly_rate = excel_row[6]  # not stored in DB
        cleaning_fee = safe_float(excel_row[7])
        gross_revenue = safe_float(excel_row[8])
        platform_fees = safe_float(excel_row[9])
        net_payout = safe_float(excel_row[10])
        notes = excel_row[11]

        if month is None and platform is None:
            continue  # truly empty row

        # Parse dates
        check_in_str = fmt_date(check_in)
        check_out_str = fmt_date(check_out)

        # For bookings without dates, synthesize from month
        if not check_in_str:
            # Try to parse month like "Mar 2026" or "April 2026"
            if month:
                month_str = str(month).strip()
                for fmt in ("%b %Y", "%B %Y"):
                    try:
                        parsed = datetime.strptime(month_str, fmt)
                        check_in_str = parsed.strftime("%Y-%m-01")
                        # Default to 1 night if unknown
                        if not check_out_str:
                            check_out_str = parsed.strftime("%Y-%m-02")
                        break
                    except ValueError:
                        continue

        if not check_in_str:
            print(f"  WARN: skipping booking with no parseable date: month={month}, guest={guest_name}")
            continue

        if not check_out_str:
            check_out_str = check_in_str  # same-day fallback

        # Extract confirmation code from notes
        confirmation_code = extract_confirmation_code(notes)
        tax_amount = extract_tax_amount(notes)

        # Platform fees: handle non-numeric (e.g., 'Active' in the direct rental row)
        if platform_fees is None:
            platform_fees = 0

        source_platform = str(platform).strip() if platform else "Unknown"

        rows.append({
            "workspace_id": WORKSPACE_ID,
            "property_id": PROPERTY_360CR,
            "source_platform": source_platform,
            "source_confirmation_code": confirmation_code,
            "guest_name": str(guest_name).strip() if guest_name else None,
            "check_in_date": check_in_str,
            "check_out_date": check_out_str,
            "nights": nights,
            "gross_revenue_amount": gross_revenue,
            "cleaning_fee_amount": cleaning_fee,
            "platform_fee_amount": platform_fees,
            "tax_amount": tax_amount,
            "net_payout_amount": net_payout,
            "status": "committed",
        })

    return rows


# ── Main ────────────────────────────────────────────────────────────────

def main():
    if not SERVICE_KEY and not DRY_RUN:
        print("ERROR: Set SUPABASE_SERVICE_KEY env var (or use --dry-run)")
        sys.exit(1)

    excel_path = os.path.abspath(EXCEL_PATH)
    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ── Expenses ──
    print("\n── Parsing Expenses sheet ──")
    expense_rows = parse_expenses(wb["Expenses"])
    biz = sum(1 for r in expense_rows if r["review_state"] == "Business")
    personal = sum(1 for r in expense_rows if r["review_state"] == "Personal")
    review = sum(1 for r in expense_rows if r["review_state"] == "Review")
    committed = sum(1 for r in expense_rows if r["status"] == "committed")
    draft = sum(1 for r in expense_rows if r["status"] == "draft")
    print(f"  Total: {len(expense_rows)} | Business: {biz} | Personal: {personal} | Review: {review}")
    print(f"  Status: {committed} committed, {draft} draft")

    # ── Bookings ──
    print("\n── Parsing Bookings sheet ──")
    booking_rows = parse_bookings(wb["Bookings"])
    print(f"  Total: {len(booking_rows)}")
    for b in booking_rows:
        print(f"    {b['check_in_date']} | {b['guest_name'] or '(no guest)'} | {b['source_platform']} | gross={b['gross_revenue_amount']} net={b['net_payout_amount']}")

    # ── Insert ──
    print("\n── Inserting into Supabase ──")
    supabase_insert("expenses", expense_rows)
    supabase_insert("bookings", booking_rows)

    print(f"\nDone! Migrated {len(expense_rows)} expenses + {len(booking_rows)} bookings.")
    if DRY_RUN:
        print("(dry run — no data was written)")


if __name__ == "__main__":
    main()
